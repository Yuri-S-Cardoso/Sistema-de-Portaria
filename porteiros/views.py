from django.shortcuts import render
from .models import Porteiro, Veiculo, Motorista, CadastroInter
from .models import CadastroInterEntrada, CadastroTerceiros, EmpresaTerceiros
from .models import CadastroTerceirosSaida, CadastroInter_Temporaria
from django.shortcuts import redirect
from django.http import JsonResponse
import json
from datetime import date, timedelta
import datetime
from django.utils import timezone
from django.db import connection
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.models import User
from .decorators import login_required_custom
from django.views.decorators.csrf import csrf_exempt
import re



#Base
def login(request):
    if request.method == "POST":
        matricula = request.POST["matricula"]
        senha = request.POST["senha"]
        try:
            porteiro = Porteiro.objects.get(matricula=matricula, senha=senha)
            request.session["porteiro_id"] = porteiro.id
            request.session["nivel"] = porteiro.nivel
            print("Nível salvo na sessão:", request.session.get("nivel"))

            return redirect("cadastro")

        except Porteiro.DoesNotExist:
            error_message = "Matrícula ou senha inválidas. Tente novamente."
            # Mantém a mesma página de login e envia a mensagem de erro
            return render(request, "porteiros/base/login.html", {"error_message": error_message})

    return render(request, "porteiros/base/login.html")

@login_required_custom
def cadastro(request):
    print("Nível do usuário na sessão:", request.session.get("nivel"))
    return render(request, "porteiros/base/cadastro.html")

@login_required_custom
def relatorio(request):
    return render(request, "porteiros/base/relatorio.html")

@login_required_custom
def cadastro_veiculo(request):
    if request.method == "POST":
        placa = request.POST["placa"]
        veiculo = request.POST["veiculo"]
        empresa = request.POST["empresa"]

        veiculos = Veiculo(placa=placa, veiculo=veiculo, empresa=empresa)
        veiculos.save()

        return redirect("cadastro")

    return render(request, "porteiros/base/veiculo.html")

@login_required_custom
def cadastro_motorista(request):
    if request.method == "POST":
        nome = request.POST["nome"]

        motoristas = Motorista(nome=nome)
        motoristas.save()
        return redirect("cadastro")

    return render(request, "porteiros/base/motorista.html")

@login_required_custom
def cadastro_usuario(request):
    if request.method == "POST":
        nome = request.POST["nome"]
        matricula = request.POST["matricula"]
        senha = request.POST["senha"]
        nivel = request.POST["nivel"]

        porteiro = Porteiro(nome=nome, matricula=matricula, senha=senha, nivel=nivel)
        porteiro.save()
        return redirect("cadastro")

    return render(request, "porteiros/base/usuario.html")

#Páginas
@login_required_custom
def cadastro_inter(request):
    if request.method == "POST":
        placa = request.POST.get("placa")
        veiculo = request.POST.get("veiculo")
        data = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        km_saida = request.POST.get("km_saida")
        lacre_saida = request.POST.get("lacre_saida")
        nfs_saida = request.POST.get("nfs_saida")
        motorista = request.POST.get("nome")
        carga = request.POST.get("carga")
        qtde_malotes_saida = request.POST.get("qtde_malotes_saida")
        carga_extra = request.POST.get("carga_extra")
        especificacao_carga = request.POST.get("especificacao_carga")

        hoje = date.today()
        entrada_existente = CadastroInter_Temporaria.objects.filter(
            placa=placa, data__date=hoje
        ).exists()

        if entrada_existente:
            return render(request, "porteiros/paginas/cadastro_inter_entrada.html")

        destino_list = request.POST.getlist("destino")
        destino = ",".join(destino_list) if destino_list else None

        outros_destinos = request.POST.get("outrosDestinos")
        if outros_destinos:
            destino = destino + "," + outros_destinos

        registro_saida = CadastroInter(
            placa=placa,
            veiculo=veiculo,
            data=data,
            km_saida=km_saida,
            lacre_saida=lacre_saida,
            nfs_saida=nfs_saida,
            motorista=motorista,
            destino=destino,
            carga=carga,
            qtde_malotes_saida=qtde_malotes_saida,
            carga_extra=carga_extra,
            especificacao_carga=especificacao_carga,
        )
        registro_saida.save()

        # Salvar também em CadastroInter_Temporaria
        registro_temporario = CadastroInter_Temporaria(
            placa=placa,
            veiculo=veiculo,
            data=data,
            km_saida=km_saida,
            lacre_saida=lacre_saida,
            nfs_saida=nfs_saida,
            motorista=motorista,
            destino=destino,
            carga=carga,
            qtde_malotes_saida=qtde_malotes_saida,
            carga_extra=carga_extra,
            especificacao_carga=especificacao_carga,
        )
        registro_temporario.save()
        
        registros_temporarios = CadastroInter_Temporaria.objects.all()

        contexto = {}  # Definindo contexto aqui
        contexto["registros_temporarios"] = registros_temporarios

        return redirect("cadastro")

    veiculos = Veiculo.objects.all()
    motoristas = Motorista.objects.all()
    agora = timezone.now().strftime("%d/%m/%Y, %H:%M:%S")
    contexto = {
        "veiculos": veiculos,
        "motoristas": motoristas,
        "data_hora_atual": agora,
    }
    return render(request, "porteiros/paginas/cadastro_inter_saida.html", contexto)

@login_required_custom
def cadastro_inter_entrada(request):
    if request.method == "POST":
        try:
            placa_entrada = request.POST.get("placa_entrada")
            km_entrada = request.POST.get("km_entrada")
            lacre_entrada = request.POST.get("lacre_entrada")
            nfs_entrada = request.POST.get("nfs_entrada")
            qtde_malotes_entrada = request.POST.get("qtde_malotes_entrada")

            deleted_count = CadastroInter_Temporaria.objects.filter(placa=placa_entrada).delete()
            #print(f"Deleted {deleted_count} record(s) from CadastroInter_Temporaria")

            data_entrada = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            registro_entrada = CadastroInterEntrada(
                placa_entrada=placa_entrada,
                data_entrada=data_entrada,
                km_entrada=km_entrada,
                lacre_entrada=lacre_entrada,
                nfs_entrada=nfs_entrada,
                qtde_malotes_entrada=qtde_malotes_entrada,
            )
            registro_entrada.save()

            return redirect("cadastro")
        except Exception as e:
            print(f"Error: {e}")

    agora = datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")
    contexto = {"data_hora_atual": agora}
    return render(request, "porteiros/paginas/cadastro_inter_entrada.html", contexto)

@login_required_custom
def cadastro_terceiros(request):
    if request.method == "POST":
        placa_entrada = request.POST.get("placa_entrada")
        veiculo_entrada = request.POST.get("veiculo_entrada")
        data = timezone.now()
        cnpj = request.POST.get("cnpj")
        nome = request.POST.get("nome")
        carga = request.POST.get("carga_entrada")
        motorista = request.POST.get("motorista")
        #documento_motorista = request.POST.get("documento_motorista")
        produto = request.POST.get("produto")        
        ajudante = request.POST.get("ajudante")
        paletes = request.POST.get("paletes")
        chapelex = request.POST.get("chapelex")
        nfs = request.POST.get("nfs")
        # nfs_prefixo = request.POST.get('nfs_prefixo')
        # nfs_de = request.POST.get('nfs_de')
        # nfs_ate = request.POST.get('nfs_ate')
        # nfs_exceto = request.POST.get('nfs_exceto')
        # nfs_apenas = request.POST.get('nfs_apenas')

        arquivo = CadastroTerceiros(
            placa_entrada=placa_entrada,
            veiculo_entrada=veiculo_entrada,
            data=data,
            cnpj=cnpj,
            nome=nome,
            carga=carga,
            motorista=motorista,
            #documento_motorista=documento_motorista,
            produto=produto,
            ajudante=ajudante,
            paletes=paletes,
            chapelex=chapelex,
            nfs=nfs
            # nfs_prefixo=nfs_prefixo,
            # nfs_de=nfs_de,
            # nfs_ate=nfs_ate,
            # nfs_exceto=nfs_exceto,
            # nfs_apenas=nfs_apenas
        )
        arquivo.save()

        agora = datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")
        if EmpresaTerceiros.objects.filter(cnpj=cnpj).exists():
            mensagem = f"O CNPJ {cnpj} já está cadastrado."
        else:
            mensagem = f"O CNPJ {cnpj} está disponível para cadastro."

        contexto = {
            "data_hora": agora,
            "mensagem": mensagem,
        }

        contexto_cupom = {
            "placa_entrada": placa_entrada,
            "veiculo_entrada": veiculo_entrada,
            "data": data,
            "cnpj": cnpj,
            "nome": nome,
            "carga": carga,
            "motorista": motorista,
            #"documento_motorista": documento_motorista,
            "produto": produto,
            "ajudante": ajudante,
            "paletes": paletes,
            "chapelex": chapelex,
            "nfs": nfs,
            #'nfs_prefixo': nfs_prefixo,
            #'nfs_de': nfs_de,
            #'nfs_ate': nfs_ate,
            #'nfs_exceto': nfs_exceto,
            #'nfs_apenas': nfs_apenas,
        }

        # Renderize o template "cupom.html" e retorne a resposta
        return render(request, "porteiros/paginas/cupom.html", contexto_cupom)

    agora = datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")
    contexto = {"data_hora_atual": agora}
    return render(request, "porteiros/paginas/cadastro_terceiros_entrada.html", contexto)

@login_required_custom
def cadastro_terceiros_saida(request):
    if request.method == "POST":
        placa_saida = request.POST.get("placa_saida")
        data_saida = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        descarga = request.POST.get("descarga")
        motivo = request.POST.get("motivo")
        descarga_paga = request.POST.get("descarga_paga")
        paletes_saida = request.POST.get("paletes_saida")
        chapelex_saida = request.POST.get("chapelex_saida")
        nfs = request.POST.get("nfs")
        # nfs_prefixo = request.POST.get('nfs_prefixo')
        # nfs_de = request.POST.get('nfs_de')
        # nfs_ate = request.POST.get('nfs_ate')
        # nfs_exceto = request.POST.get('nfs_exceto')
        # nfs_apenas = request.POST.get('nfs_apenas')

        arquivo = CadastroTerceirosSaida(
            placa_saida=placa_saida,
            data_saida=data_saida,
            descarga=descarga,
            motivo=motivo,
            descarga_paga=descarga_paga,
            paletes_saida=paletes_saida,
            chapelex_saida=chapelex_saida,
            nfs=nfs
            # nfs_prefixo=nfs_prefixo,
            # nfs_de=nfs_de,
            # nfs_ate=nfs_ate,
            # nfs_exceto=nfs_exceto,
            # nfs_apenas=nfs_apenas
        )
        arquivo.save()

        return redirect("cadastro")

    ultimo_cadastro = CadastroTerceiros.objects.last()
    contexto = {"cadastro": ultimo_cadastro}
    return render(request, "porteiros/paginas/cadastro_terceiros_saida.html", contexto)

@login_required_custom
def cupom(request):
    return render(request, "porteiros/paginas/cupom.html")

@login_required_custom
def relatorio_inter_entrada(request):
    selected_data_inicio = request.GET.get(
        "data_inicio", date.today().strftime("%Y-%m-%d")
    )
    selected_data_fim = request.GET.get("data_fim", date.today().strftime("%Y-%m-%d"))
    selected_placa = request.GET.get("placa")

    data = CadastroInterEntrada.objects.all()

    if selected_data_inicio:
        data = data.filter(data_entrada__date__gte=selected_data_inicio)

    if selected_data_fim:
        data = data.filter(data_entrada__date__lte=selected_data_fim)

    if selected_placa:
        data = data.filter(placa_entrada=selected_placa)

    placas = (
        CadastroInterEntrada.objects.filter(
            data_entrada__date__range=[selected_data_inicio, selected_data_fim]
        )
        .values_list("placa_entrada", flat=True)
        .distinct()
    )
    
    if request.GET.get('export') == 'excel':
        return export_to_excel_entrada(data)

    return render(
        request,
        "porteiros/paginas/relatorio_inter_entrada.html",
        {
            "placas": placas,
            "selected_data_inicio": selected_data_inicio,
            "selected_data_fim": selected_data_fim,
            "selected_placa": selected_placa,
            "data": data,
        },
    )

@login_required_custom    
def relatorio_inter_saida(request):
    
    
    selected_data_inicio = request.GET.get("data_inicio", date.today())
    selected_data_fim = request.GET.get("data_fim", date.today())
    selected_placa = request.GET.get("placa")

    data = CadastroInter.objects.all()

    if selected_data_inicio:
        data = data.filter(data__date__gte=selected_data_inicio)

    if selected_data_fim:
        data = data.filter(data__date__lte=selected_data_fim)

    if selected_placa:
        data = data.filter(placa=selected_placa)

    placas = (
        CadastroInter.objects.filter(
            data__date__range=[selected_data_inicio, selected_data_fim]
        )
        .values_list("placa", flat=True)
        .distinct()
    )
    
    if request.GET.get('export') == 'excel':
        return export_to_excel_saida(data)

    return render(
        request,
        "porteiros/paginas/relatorio_inter_saida.html",
        {
            "placas": placas,
            "selected_data_inicio": selected_data_inicio,
            "selected_data_fim": selected_data_fim,
            "selected_placa": selected_placa,
            "data": data,
        },
    )
    
@login_required_custom    
def relatorio_terceiros_entrada(request):
    selected_data_inicio = request.GET.get(
        "data_inicio", date.today().strftime("%Y-%m-%d")
    )
    selected_data_fim = request.GET.get("data_fim", date.today().strftime("%Y-%m-%d"))
    selected_placa = request.GET.get("placa")

    data = CadastroTerceiros.objects.all()

    if selected_data_inicio:
        data = data.filter(data__date__gte=selected_data_inicio)

    if selected_data_fim:
        data = data.filter(data__date__lte=selected_data_fim)

    if selected_placa:
        data = data.filter(placa_entrada=selected_placa)

    placas = (
        CadastroTerceiros.objects.filter(
            data__date__range=[selected_data_inicio, selected_data_fim]
        )
        .values_list("placa_entrada", flat=True)
        .distinct()
    )

    return render(
        request,
        "porteiros/paginas/relatorio_terceiros_entrada.html",
        {
            "placas": placas,
            "selected_data_inicio": selected_data_inicio,
            "selected_data_fim": selected_data_fim,
            "selected_placa": selected_placa,
            "data": data,
        },
    )

@login_required_custom
def relatorio_terceiros_saida(request):
    selected_data_inicio = request.GET.get(
        "data_inicio", date.today().strftime("%Y-%m-%d")
    )
    selected_data_fim = request.GET.get("data_fim", date.today().strftime("%Y-%m-%d"))
    selected_placa = request.GET.get("placa")

    data = CadastroTerceirosSaida.objects.all()

    if selected_data_inicio:
        data = data.filter(data_saida__date__gte=selected_data_inicio)

    if selected_data_fim:
        data = data.filter(data_saida__date__lte=selected_data_fim)

    if selected_placa:
        data = data.filter(placa_saida=selected_placa)

    placas = (
        CadastroTerceirosSaida.objects.filter(
            data_saida__date__range=[selected_data_inicio, selected_data_fim]
        )
        .values_list("placa_saida", flat=True)
        .distinct()
    )

    return render(
        request,
        "porteiros/paginas/relatorio_terceiros_saida.html",
        {
            "placas": placas,
            "selected_data_inicio": selected_data_inicio,
            "selected_data_fim": selected_data_fim,
            "selected_placa": selected_placa,
            "data": data,
        },
    )    
    
#Configurações    
def verificar_entrada(request):
    if request.method == "POST":
        placa = request.POST.get("placa")
        hoje = date.today()
        entrada_cadastro_inter_temporaria = CadastroInter_Temporaria.objects.filter(
            placa=placa, data__date=hoje
        ).exists()
        print("placa", placa)
        print("existe_entrada_temporaria:", entrada_cadastro_inter_temporaria)
        
        return JsonResponse(
            {
                "existe_entrada_temporaria": entrada_cadastro_inter_temporaria,
            }
        )

def buscar_veiculo(request):
    placa = request.POST.get("placa")
    veiculo = Veiculo.objects.filter(placa=placa).first()

    if veiculo:
        response = {
            "veiculo": veiculo.veiculo,
        }
    else:
        response = {
            "veiculo": "",
        }

    return JsonResponse(response)

def cadastro_cnpj_terceiros(request):
    if request.method == "POST":
        nome = request.POST.get("nome")
        cnpj = request.POST.get("cnpj")

        empresa = EmpresaTerceiros(
            nome=nome,
            cnpj=cnpj,
        )
        empresa.save()

        return redirect("cadastro_terceiros")

    return render(request, "porteiros/paginas/cadastro_cnpj_terceiros.html")

def verificar_cnpj(request):
    if request.method == "POST":
        cnpj = request.POST.get("cnpj", "").strip()
        
        # Limpar o CNPJ (remover pontos, traços e barras)
        cnpj_limpo = re.sub(r'[^0-9]', '', cnpj)
        
        try:
            # Verificar se o CNPJ existe no model EmpresaTerceiros
            empresa = EmpresaTerceiros.objects.get(cnpj=cnpj_limpo)
            return JsonResponse({"existe": True, "nome_empresa": empresa.nome})
        except EmpresaTerceiros.DoesNotExist:
            return JsonResponse({"existe": False})
        except ValueError:
            return JsonResponse({"erro": "CNPJ inválido"}, status=400)
        
def buscar_nome_por_cnpj(request):
    if request.method == "POST":
        cnpj = request.POST.get("cnpj")
        
        # Consulta SQL para buscar o nome correspondente ao CNPJ na tabela rms.view_fornecedor
        with connection.cursor() as cursor:
            cursor.execute("SELECT tip_razao_social FROM view_fornecedor WHERE tip_cgc_cpf = %s", [cnpj])
            razao_social = cursor.fetchone()

        if razao_social:
            nome = razao_social[0]  # A razão social está na primeira (e única) coluna
            return JsonResponse({"nome": nome})
        else:
            return JsonResponse({"error": "Empresa não encontrada para o CNPJ fornecido: {}".format(cnpj)})

def verificar_placa(request):
    if request.method == "POST":
        placa = request.POST.get("placa")
        data_atual = timezone.now().date()

        try:
            registro = CadastroTerceiros.objects.get(
                placa_entrada=placa, data=data_atual
            )
            return JsonResponse(
                {"placa_existente": True, "registro": registro.placa_entrada}
            )
        except CadastroTerceiros.DoesNotExist:
            return JsonResponse({"placa_existente": False})

    return JsonResponse({"placa_existente": False})

def export_to_excel_saida(data):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=relatorio_inter_saida.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Relatório Inter Saída'

    columns = [
        'Placa', 'Veículo', 'Data', 'Km Saída', 'Lacre Saída', 'NFS Saída', 'Motorista',
        'Destino', 'Carga', 'Qtde Malotes Saída', 'Carga Extra', 'Especificação Carga'
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for item in data:
        row_num += 1
        row = [
            item.placa, item.veiculo, item.data.strftime("%d/%m/%Y %H:%M"), item.km_saida,
            item.lacre_saida, item.nfs_saida, item.motorista, item.destino, item.carga,
            item.qtde_malotes_saida, item.carga_extra, item.especificacao_carga
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(response)
    return response

def export_to_excel_entrada(data):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=relatorio_inter_entrada.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Relatório Inter Entrada'

    columns = [
        'Placa', 'Data', 'Km Entrada', 'Lacre Entrada', 'NFS Entrada', 'Qtde Malotes Entrada'
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for item in data:
        row_num += 1
        row = [
            item.placa_entrada, item.data_entrada.strftime("%d/%m/%Y %H:%M"), item.km_entrada,
            item.lacre_entrada, item.nfs_entrada, item.qtde_malotes_entrada
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(response)
    return response

def busca_terceiros(request):
    if request.method == "POST":
        data = json.loads(request.body)
        placa_entrada = data.get("placa", "").strip().upper()
        current_date = timezone.now().date()

        # Calcula a data limite, 1 dia antes da data atual e 2 dias depois
        start_date = current_date - timedelta(days=1)
        end_date = current_date + timedelta(days=2)

        # Consulta registros de entrada dentro do intervalo de data
        entrada_existente = CadastroTerceiros.objects.filter(
            placa_entrada=placa_entrada,
            data__date__range=[start_date, end_date] # intervalo de datas
        ).exists()

        # Consulta registros de saída apenas na data atual
        saida_existente = CadastroTerceirosSaida.objects.filter(
            placa_saida=placa_entrada,
            data_saida__date=current_date  #data atual
        ).exists()

        data = {
            "placa_entrada_existente": entrada_existente,
            "placa_saida_existente": saida_existente,
        }
        return JsonResponse(data)

    return JsonResponse({})

def logout(request):
    request.session.flush()  # Remove todos os dados da sessão
    return redirect("login")

def verificar_sessao(request):
    if not request.session.get('nivel'):  # Se não houver sessão ativa
        return redirect('login')  # Redireciona para a página de login